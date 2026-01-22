"""
Application Streamlit de gestion des inscriptions au Focus Time.

Cette application permet :
- aux √©l√®ves de s‚Äôinscrire √† des rem√©diations ou ateliers (P9, P10 ou P9+P10),
- aux professeurs de consulter les groupes et d‚Äôinscrire manuellement des √©l√®ves,
- de g√©rer les places disponibles par activit√© et par degr√©,
- de stocker les inscriptions dans une base Supabase.

Le comportement de l‚Äôapplication d√©pend :
- du r√¥le de l‚Äôutilisateur (√©l√®ve ou professeur),
- de la p√©riode d‚Äôinscription ouverte,
- du mode ATELIER_MODE (True = Inscriptions pour des ateliers, False = Inscriptions pour des rem√©diations).

Donn√©es externes :
(Permet de d√©finir le nom des options et le nombre place maximum).
- options.json : activit√©s P9 et P10
- options_p910.json : activit√©s qui durent deux p√©riodes
- registration_open.json : fen√™tre temporelle d‚Äôinscription
"""


import streamlit as st
import json
import httpx
import string
from io import BytesIO
from openpyxl import Workbook
from supabase import create_client, Client
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment

TIMEZONE = 1  # GMT+1
DEGREE_PROF = 4

ATELIER_MODE = True


@st.cache_resource
def init_db_connection() -> Client:
    """
        Initialise et met en cache la connexion √† la base de donn√©es Supabase.

        Les identifiants sont r√©cup√©r√©s depuis les secrets Streamlit :
        - SUPABASE_URL
        - SUPABASE_KEY

        Returns:
            Client: Client Supabase pr√™t √† √™tre utilis√© pour les requ√™tes
                    (tables students, options, etc.).
    """

    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]

    client = create_client(url, key)
    # user = client.auth.sign_in_with_password({"email": st.secrets["USER_EMAIL"], "password": st.secrets["USER_PASS"]})

    return client


@st.dialog("Inscrire un √©l√®ve", width="medium")
def select_student():
    """
        Interface professeur pour inscrire manuellement un √©l√®ve.

        - r√©cup√®re la liste des √©l√®ves depuis la table `students`,
        - v√©rifie si l‚Äô√©l√®ve est d√©j√† inscrit en P9, P10 ou P910,
        - emp√™che les doubles inscriptions et les groupes complets,
        - ins√®re les nouvelles inscriptions dans la table `options`.
    """

    try:
        response_email = client.table("students").select("*").execute()
    except httpx.ReadError:
        st.rerun()

    all_emails = []
    for email in response_email.data:
        all_emails.append(email["email"].lower())

    email = st.selectbox(
        "Adresse email de l'√©l√®ve",
        all_emails,
        index=None,
        placeholder="Choisir un email"
    )

    enroll_p9 = False
    enroll_p10 = False
    if email is not None:
        for enroll in response_options.data:
            if enroll["email"].lower() == email.lower():
                if int(enroll['period']) == 9:
                    enroll_p9 = True
                elif int(enroll['period']) == 10:
                    enroll_p10 = True
                elif int(enroll['period']) == 910:
                    enroll_p9 = True
                    enroll_p10 = True
                st.success(f"{enroll['name']} est d√©j√† inscrit en {enroll['choice']} (P{enroll['period']})")

    opts_list = []
    for degree, option_names in options_list.items():
        for name in option_names.keys():
            opts_list.append(f"{name} ({degree})")

    opts_list_p910 = []
    for degree, option_names in options_p910_list.items():
        for name in option_names.keys():
            opts_list_p910.append(f"{name} ({degree})")

    option_p9 = st.selectbox(
        "Rem√©diation/Atelier P9",
        opts_list,
        index=None,
        placeholder="Choisir une rem√©diation/atelier"
    )

    no_place_left = False
    if option_p9 is not None:
        option = " ".join(option_p9.split()[:-1]) + f" P9 D{option_p9.split()[-1][2]}"

        place_total = options_list[f"D{option_p9.split()[-1][2]}"][" ".join(option_p9.split()[:-1])]

        if option in already_registered:
            place_left = place_total - already_registered[option]
            if place_left > 0:
                st.info(f"{already_registered[option]} √©l√®ves d√©j√† inscrits en {option_p9} (P9)")
            else:
                no_place_left = True
                st.error("Le groupe est complet")

        if enroll_p9:
            no_place_left = True
            st.error("Cet √©l√®ve a d√©j√† une inscription en P9")

    option_p10 = st.selectbox(
        "Rem√©diation/Atelier P10",
        opts_list,
        index=None,
        placeholder="Choisir une rem√©diation/atelier"
    )

    if option_p10 is not None:
        option = " ".join(option_p10.split()[:-1]) + f" P10 D{option_p10.split()[-1][2]}"

        place_total = options_list[f"D{option_p10.split()[-1][2]}"][" ".join(option_p10.split()[:-1])]

        if option in already_registered:
            place_left = place_total - already_registered[option]
            if place_left > 0:
                st.info(f"{already_registered[option]} √©l√®ves d√©j√† inscrits en {option_p10} (P10)")
            else:
                no_place_left = True
                st.error("Le groupe est complet")

        if enroll_p10:
            no_place_left = True
            st.error("Cet √©l√®ve a d√©j√† une inscription en P10")

    option_p910 = st.selectbox(
        "Rem√©diation/Atelier P9 et P10",
        opts_list_p910,
        index=None,
        placeholder="Choisir une rem√©diation/atelier"
    )

    if option_p910 is not None:
        option = " ".join(option_p910.split()[:-1]) + f" P910 D{option_p910.split()[-1][2]}"
        place_total = options_p910_list[f"D{option_p910.split()[-1][2]}"][" ".join(option_p910.split()[:-1])]

        if option in already_registered:
            place_left = place_total - already_registered[option]
            if place_total == 0 or place_left > 0:
                st.info(f"{already_registered[option]} √©l√®ves d√©j√† inscrits en {option_p910} (P9 et P10)")
            else:
                no_place_left = True
                st.error("Le groupe est complet")

        if enroll_p9 or enroll_p10:
            no_place_left = True
            st.error("Cet √©l√®ve a d√©j√† une inscription en P9 ou P10")

    st.divider()
    if st.button("Valider", disabled=no_place_left):
        if email is not None:
            name = email.split("@")[0].split(".")
            if len(name) > 1:
                name = name[0].capitalize() + " " + name[1].capitalize()
            else:
                name = name[0].capitalize()
            if option_p9 is not None:
                data = {
                    "email": email,
                    "name": name,
                    "choice": " ".join(option_p9.split()[:-1]),
                    "period": 9,
                    "degree": int(option_p9.split()[-1][2])
                }
                client.table("options").insert(data).execute()
            if option_p10 is not None:
                data = {
                    "email": email,
                    "name": name,
                    "choice": " ".join(option_p10.split()[:-1]),
                    "period": 10,
                    "degree": int(option_p10.split()[-1][2])
                }
                client.table("options").insert(data).execute()
            if option_p910 is not None:
                data = {
                    "email": email,
                    "name": name,
                    "choice": " ".join(option_p910.split()[:-1]),
                    "period": 910,
                    "degree": int(option_p910.split()[-1][2])
                }
                client.table("options").insert(data).execute()
            if option_p9 is not None or option_p10 is not None or option_p910 is not None:
                st.rerun()


def gen_form(title, period, place):
    """
        G√©n√®re un formulaire Streamlit pour une activit√© donn√©e.

        Le formulaire affiche :
        - le nom de l‚Äôactivit√©,
        - le nombre de places restantes,
        - un bouton d‚Äôinscription activ√© ou non selon la disponibilit√©.

        En cas de soumission valide, l‚Äôinscription est enregistr√©e
        dans la table `options`.

        Args:
            title (str): Nom de l‚Äôactivit√©.
            period (int): P√©riode concern√©e (9, 10 ou 910).
            place (int): Nombre de places restantes.
    """

    with st.form(title + f"_p{period}"):
        st.write(title)

        col1, col2 = st.columns([3, 1])
        with col1:
            if place > 3:
                st.info(f"Il reste {place} places")
            elif place > 0:
                if place > 1:
                    st.warning(f"Il ne reste plus que {place} places")
                else:
                    st.warning(f"Il ne reste plus que {place} place")
            else:
                st.error("Il n'y a plus de place")
        with col2:
            if place > 0:
                submitted = st.form_submit_button("S'inscrire", width="stretch")
            else:
                submitted = st.form_submit_button("S'inscrire", width="stretch", disabled=True)
        if submitted:
            data = {
                "email": student_email,
                "name": student_name,
                "choice": title,
                "period": period,
                "degree": student_degree
            }
            if place > 0:
                client.table("options").insert(data).execute()
            st.rerun()

def gen_registration(period: int):
    """
        G√©n√®re l‚Äôensemble des formulaires d‚Äôinscription pour une p√©riode donn√©e.

        Selon la p√©riode, la fonction :
        - affiche les rem√©diations ou ateliers correspondants,
        - filtre les options selon le degr√© de l‚Äô√©l√®ve,
        - tient compte des inscriptions d√©j√† existantes,
        - g√®re les groupes communs D2/D3.

        Args:
            period (int):
                - 9   : rem√©diations / ateliers P9
                - 10  : rem√©diations / ateliers P10
                - 910 : rem√©diations / ateliers P9 et P10 combin√©s
    """

    if period == 910:
        # P9 ET P10
        if ATELIER_MODE:
            st.markdown(f"#### Ateliers")
        else:
            st.markdown(f"#### Rem√©diations P9 et P10")
        choice_list = options_p910_list
    else:
        if ATELIER_MODE:
            st.markdown(f"#### Ateliers P{period}")
        else:
            st.markdown(f"#### Rem√©diations P{period}")
        choice_list = options_list

    for title, place in choice_list[f"D{student_degree}"].items():
        if title + f" P{period} D{student_degree}" in already_registered:
            place = max(place - already_registered[title + f" P{period} D{student_degree}"], 0)
        gen_form(title, period, place)
    if "D2_D3" in choice_list and student_degree > 1:
        for title, place in choice_list["D2_D3"].items():
            if title + f" P{period} D2" in already_registered:
                place = max(place - already_registered[title + f" P{period} D2"], 0)
            if title + f" P{period} D3" in already_registered:
                place = max(place - already_registered[title + f" P{period} D3"], 0)
            gen_form(title, period, place)

    st.divider()


def get_not_registered():
    try:
        all_students = client.table("students").select("*").execute()
    except httpx.ReadError:
        st.rerun()

    all_email_d1 = []
    all_email_d2 = []
    all_email_d3 = []

    for student in all_students.data:
        email = student["email"].lower()
        if student["degree"] == 1:
            all_email_d1.append(email)
        elif student["degree"] == 2:
            all_email_d2.append(email)
        elif student["degree"] == 3:
            all_email_d3.append(email)

    for student_registered in response_options.data:
        degree = student_registered["degree"]
        email = student_registered["email"].lower()

        if degree == 1 and email in all_email_d1:
            all_email_d1.remove(email)
        if degree == 2 and email in all_email_d2:
            all_email_d2.remove(email)
        if degree == 3 and email in all_email_d3:
            all_email_d3.remove(email)

    return all_email_d1, all_email_d2, all_email_d3


def create_excel_file():
    wb = Workbook()
    ws = wb.active

    ws.title = "D1"
    wb.create_sheet("D2")
    wb.create_sheet("D3")
    wb.create_sheet("D2-D3")

    alphabetic = string.ascii_uppercase
    colors = ["FF99CC", "CC99FF", "FFCC99", "3366FF", "33CCCC"]

    for sheet, degree in enumerate(["D1", "D2", "D3", "D2_D3"]):
        wb.active = sheet
        ws = wb.active
        ws.row_dimensions[1].height = 50
        # Set title
        for index, option_name in enumerate(options_p910_list[degree]):
            set_color = colors[index % len(colors)]
            ws[f"{alphabetic[index]}1"] = option_name
            ws[f"{alphabetic[index]}1"].fill = PatternFill(start_color=set_color, end_color=set_color, fill_type="solid")
            ws[f"{alphabetic[index]}1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[f"{alphabetic[index]}"].width = 40

            row = 2
            for data in response_options.data:
                if data["choice"] == option_name:
                    first_name = data["email"].split(".")[0].lower()
                    name = data["email"].split("@")[0].split(".")[1].lower()
                    ws[f"{alphabetic[index]}{row}"] = name.title() + " " + first_name.capitalize()
                    row += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer


st.set_page_config(page_title="Focus Time", page_icon="üìö", initial_sidebar_state="auto")
st.title("Focus Time")
st.sidebar.text("Plateforme d'inscription aux activit√©s du Focus Time")
st.sidebar.image("isa_icon.jpg")

# if False:
if not st.user.is_logged_in:
    st.warning("Connecte toi avant de continuer")
    if st.button("Connexion"):
        st.login("microsoft")
else:
    student_name = st.user.name
    student_email = st.user.email
    # student_name = "Test1"
    # student_email = "test1@isa-florenville.be"
    student_degree = 0  # 0 = not fetched yet, 4 = Prof
    registered_options = []

    rem_p9 = False
    rem_p10 = False

    with open("options.json", "r", encoding="utf-8") as file:
        options_list = json.load(file)
    with open("options_p910.json", "r", encoding="utf-8") as file:
        options_p910_list = json.load(file)

    client = init_db_connection()
    try:
        response_degree = client.table("students").select("degree").ilike("email", student_email).execute()
    except httpx.ReadError:
        st.rerun()

    if len(response_degree.data) > 0:
        student_degree = int(response_degree.data[0]["degree"])

    try:
        response_options = client.table("options").select("*").execute()
    except httpx.ReadError:
        st.rerun()

    already_registered = {}
    for data in response_options.data:
        if data["email"].lower() == student_email.lower():
            registered_options.append(data)

        key = data["choice"] + f" P{data['period']} D{data['degree']}"
        if key in already_registered:
            already_registered[key] += 1
        else:
            already_registered[key] = 1

    # region Sidebar
    st.sidebar.divider()
    if student_degree == DEGREE_PROF:
        st.sidebar.write(f"Bonjour, {student_name} ! (Prof)")
    else:
        st.sidebar.write(f"Bonjour, {student_name} ! (D{student_degree})")

    if st.sidebar.button("D√©connexion"):
        st.logout()
    # endregion

    if student_degree == DEGREE_PROF:

        if st.button("Inscrire un √©l√®ve", width="stretch", type="primary"):
            select_student()
        if st.button("Voir les groupes", width="stretch"):
            get_not_registered()
            if len(response_options.data) > 0:
                with st.container(border=True):
                    table_data = {}
                    for data in response_options.data:
                        if not data["choice"] in table_data:
                            table_data[data["choice"]] = []
                        table_data[data["choice"]].append({"name": data["name"],
                                                           "degree": data["degree"],
                                                           "period": data["period"]})

                    for key, value in table_data.items():
                        st.markdown(f"**{key}**")

                        st.dataframe(value, use_container_width=True, hide_index=True,
                                     column_order=["name", "degree", "period"],
                                     column_config={"name": "Pr√©nom/Nom",
                                                    "degree": st.column_config.NumberColumn(
                                                        "Degr√©",
                                                        format="D%d",
                                                    ),
                                                    "period": st.column_config.NumberColumn(
                                                        "P√©riode",
                                                        format="P%d",
                                                    )})
                        st.divider()
                with st.expander("Pas inscrit"):
                    not_reg = get_not_registered()
                    not_reg_d1 = [" ".join(name.split("@")[0].split(".")).title() for name in not_reg[0]]
                    not_reg_d2 = [" ".join(name.split("@")[0].split(".")).title() for name in not_reg[1]]
                    not_reg_d3 = [" ".join(name.split("@")[0].split(".")).title() for name in not_reg[2]]
                    with st.expander("D1"):
                        st.write(f"{len(not_reg_d1)} √©l√®ves ne sont pas inscrit")
                        st.dataframe(not_reg_d1, column_config={"value": "Pr√©nom/Nom"})
                    with st.expander("D2"):
                        st.write(f"{len(not_reg_d2)} √©l√®ves ne sont pas inscrit")
                        st.dataframe(not_reg_d2, column_config={"value": "Pr√©nom/Nom"})
                    with st.expander("D3"):
                        st.write(f"{len(not_reg_d3)} √©l√®ves ne sont pas inscrit")
                        st.dataframe(not_reg_d3, column_config={"value": "Pr√©nom/Nom"})

                st.download_button("Exporter en fichier Excel", width="stretch", type="primary",
                                   data=create_excel_file,
                                   file_name="export.xlsx",
                                   on_click="ignore")
            else:
                st.info("Aucun groupe pour l'instant")
    else:
        with open("registration_open.json", "r", encoding="utf-8") as file:
            regis_open = json.load(file)

        target_time = datetime.strptime(regis_open["from"] + " " + regis_open["from_hour"],
                                        "%d/%m/%Y %Hh%M")
        target_time = target_time.replace(hour=target_time.hour - TIMEZONE)

        close_date = datetime.strptime(regis_open["for"],"%d/%m/%Y").date()

        today = datetime.today().date()
        now = datetime.now()
        days_diff = (target_time.date() - today).days

        registration_open = False
        if now < target_time or today >= close_date:
            st.info("Aucune inscription pour le moment üòä")
            if 0 <= days_diff <= 3:
                st.info(f"Prochaine inscription le {regis_open['from']} √† {regis_open['from_hour']}")
            st.divider()
        else:
            registration_open = True

        if len(registered_options) > 0:
            st.text(f"Pour le {regis_open['for']} :")
            for choice in registered_options:
                if choice["period"] == 910:
                    st.success(f"Tu es inscrit en {choice["choice"]} (P9 et P10)")
                else:
                    st.success(f"Tu es inscrit en {choice["choice"]} (P{choice["period"]})")

                if choice["period"] == 9:
                    rem_p9 = True
                elif choice["period"] == 10:
                    rem_p10 = True
                elif choice["period"] == 910:
                    rem_p9 = True
                    rem_p10 = True
            st.divider()

        if registration_open:
            no_registration = True
            if student_degree >= 1:
                if len(options_list[f"D{student_degree}"]) > 0:
                    if not rem_p9:
                        gen_registration(period=9)
                    if not rem_p10:
                        gen_registration(period=10)
                    no_registration = False
                if len(options_p910_list[f"D{student_degree}"]) > 0 or (len(options_p910_list["D2_D3"]) and student_degree >= 2):
                    if not rem_p9 and not rem_p10:
                        gen_registration(period=910)
                    no_registration = False

            if no_registration:
                st.info("Aucune inscription pour toi")

